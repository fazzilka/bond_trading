import hashlib
import io
import tempfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
from numbers_parser import Document
from python_calamine import CalamineWorkbook

from bond_trading.domain.errors import InvalidIsinError
from bond_trading.domain.value_objects import normalize_isin

DEFAULT_SHEET_NAME = "Доход счёт 2026"
OOXML_EXTENSIONS = frozenset({".xlsx", ".xlsm", ".xltx", ".xltm"})
CALAMINE_EXTENSIONS = frozenset({".xls", ".xlsb"})
NUMBERS_EXTENSIONS = frozenset({".numbers"})
SUPPORTED_EXTENSIONS = OOXML_EXTENSIONS | CALAMINE_EXTENSIONS | NUMBERS_EXTENSIONS
ALLOWED_MEDIA_TYPES = frozenset(
    {
        "application/octet-stream",
        "application/vnd.apple.numbers",
        "application/vnd.ms-excel",
        "application/vnd.ms-excel.sheet.binary.macroenabled.12",
        "application/vnd.ms-excel.sheet.macroenabled.12",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/x-iwork-numbers-sffnumbers",
        "application/x-ole-storage",
        "application/x-zip-compressed",
        "application/zip",
    }
)


class SpreadsheetImportError(ValueError):
    pass


XlsxImportError = SpreadsheetImportError


def validate_spreadsheet_upload(file_name: str, content_type: str | None) -> None:
    suffix = Path(file_name).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        extensions = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise SpreadsheetImportError(f"Supported spreadsheet extensions: {extensions}")
    normalized_content_type = (content_type or "").partition(";")[0].strip().lower()
    if normalized_content_type and normalized_content_type not in ALLOWED_MEDIA_TYPES:
        raise SpreadsheetImportError(
            f"The upload MIME type {normalized_content_type!r} is not valid for {suffix}"
        )


validate_xlsx_upload = validate_spreadsheet_upload


@dataclass(frozen=True, slots=True)
class SpreadsheetCell:
    value: Any
    is_formula: bool = False


@dataclass(frozen=True, slots=True)
class SpreadsheetSheet:
    title: str
    rows: tuple[tuple[SpreadsheetCell, ...], ...]

    @property
    def max_row(self) -> int:
        return len(self.rows)

    def cell(self, row: int, column: int) -> SpreadsheetCell:
        if row <= 0 or column <= 0 or row > len(self.rows):
            return SpreadsheetCell(None)
        values = self.rows[row - 1]
        if column > len(values):
            return SpreadsheetCell(None)
        return values[column - 1]


@dataclass(frozen=True, slots=True)
class ImportRowError:
    row_number: int
    field: str
    message: str
    source_value: str | None = None


@dataclass(frozen=True, slots=True)
class ImportRow:
    row_number: int
    source_isin: str
    normalized_isin: str
    source_name: str | None
    target_event_type: str
    purchase_date: date
    target_event_date: date
    quantity: Decimal
    purchase_clean_price_rub_per_bond: Decimal
    purchase_accrued_interest_rub_per_bond: Decimal
    purchase_commission_rub_per_bond: Decimal
    target_redemption_price_rub_per_bond: Decimal | None
    sale_commission_rub_per_bond: Decimal
    planned_yield_manual_reference: Decimal | None
    offer_submission_period: str | None


@dataclass(frozen=True, slots=True)
class ImportPreview:
    file_name: str
    sheet_name: str
    checksum: str
    header_row_number: int
    rows_read: int
    rows: tuple[ImportRow, ...]
    errors: tuple[ImportRowError, ...]
    upload_id: UUID | None = None


class SpreadsheetPortfolioReader:
    def preview(
        self,
        content: bytes,
        file_name: str,
        *,
        sheet_name: str = DEFAULT_SHEET_NAME,
    ) -> ImportPreview:
        validate_spreadsheet_upload(file_name, None)
        sheets = self._load_sheets(content, file_name)
        worksheet = self._select_sheet(sheets, sheet_name)
        header_row, isin_column = self._find_header(worksheet)
        rows: list[ImportRow] = []
        errors: list[ImportRowError] = []
        rows_read = 0
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            cells = worksheet.rows[row_number - 1]
            if self._is_empty_row(cells):
                break
            rows_read += 1
            try:
                rows.append(self._parse_row(worksheet, row_number, isin_column))
            except _RowValidationError as exc:
                errors.append(
                    ImportRowError(
                        row_number=row_number,
                        field=exc.field,
                        message=exc.message,
                        source_value=exc.source_value,
                    )
                )
        return ImportPreview(
            file_name=Path(file_name).name,
            sheet_name=worksheet.title,
            checksum=hashlib.sha256(content).hexdigest(),
            header_row_number=header_row,
            rows_read=rows_read,
            rows=tuple(rows),
            errors=tuple(errors),
        )

    def _load_sheets(self, content: bytes, file_name: str) -> tuple[SpreadsheetSheet, ...]:
        suffix = Path(file_name).suffix.lower()
        try:
            if suffix in OOXML_EXTENSIONS:
                return self._load_ooxml(content, suffix)
            if suffix in CALAMINE_EXTENSIONS:
                return self._load_calamine(content)
            if suffix in NUMBERS_EXTENSIONS:
                return self._load_numbers(content)
        except SpreadsheetImportError:
            raise
        except Exception as exc:
            raise SpreadsheetImportError(
                f"The uploaded {suffix} file is not a readable spreadsheet"
            ) from exc
        raise SpreadsheetImportError(f"Unsupported spreadsheet extension: {suffix}")

    @staticmethod
    def _load_ooxml(content: bytes, suffix: str) -> tuple[SpreadsheetSheet, ...]:
        workbook = openpyxl.load_workbook(
            io.BytesIO(content),
            data_only=False,
            read_only=False,
            keep_vba=suffix in {".xlsm", ".xltm"},
        )
        return tuple(
            SpreadsheetSheet(
                title=worksheet.title,
                rows=tuple(
                    tuple(
                        SpreadsheetCell(
                            value=cell.value,
                            is_formula=cell.data_type == "f"
                            or (isinstance(cell.value, str) and cell.value.startswith("=")),
                        )
                        for cell in row
                    )
                    for row in worksheet.iter_rows()
                ),
            )
            for worksheet in workbook.worksheets
        )

    @staticmethod
    def _load_calamine(content: bytes) -> tuple[SpreadsheetSheet, ...]:
        workbook = CalamineWorkbook.from_filelike(io.BytesIO(content))
        return tuple(
            SpreadsheetSheet(
                title=name,
                rows=tuple(
                    tuple(SpreadsheetCell(value=value) for value in row)
                    for row in workbook.get_sheet_by_name(name).to_python(skip_empty_area=False)
                ),
            )
            for name in workbook.sheet_names
        )

    @staticmethod
    def _load_numbers(content: bytes) -> tuple[SpreadsheetSheet, ...]:
        with tempfile.NamedTemporaryFile(suffix=".numbers") as source:
            source.write(content)
            source.flush()
            document = Document(source.name)
            sheets: list[SpreadsheetSheet] = []
            for sheet in document.sheets:
                for table_index, table in enumerate(sheet.tables):
                    title = sheet.name if table_index == 0 else f"{sheet.name} — {table.name}"
                    sheets.append(
                        SpreadsheetSheet(
                            title=title,
                            rows=tuple(
                                tuple(
                                    SpreadsheetCell(
                                        value=cell.value,
                                        is_formula=bool(getattr(cell, "is_formula", False)),
                                    )
                                    for cell in row
                                )
                                for row in table.rows()
                            ),
                        )
                    )
            return tuple(sheets)

    def _select_sheet(
        self, worksheets: Sequence[SpreadsheetSheet], requested: str
    ) -> SpreadsheetSheet:
        requested_sheets = [
            sheet
            for sheet in worksheets
            if sheet.title.strip() == requested or sheet.title.strip().startswith(f"{requested} — ")
        ]
        for worksheet in (*requested_sheets, *worksheets):
            try:
                self._find_header(worksheet)
            except SpreadsheetImportError:
                continue
            return worksheet
        raise SpreadsheetImportError(
            f"Sheet {requested!r} or another sheet with an ISIN header was not found"
        )

    @staticmethod
    def _find_header(worksheet: SpreadsheetSheet) -> tuple[int, int]:
        for row_index, row in enumerate(worksheet.rows, start=1):
            for column_index, cell in enumerate(row, start=1):
                if str(cell.value or "").strip().upper() == "ISIN":
                    return row_index, column_index
        raise SpreadsheetImportError(f"ISIN header was not found on sheet {worksheet.title!r}")

    def _parse_row(self, worksheet: SpreadsheetSheet, row: int, isin_column: int) -> ImportRow:
        def cell(column: int) -> SpreadsheetCell:
            return worksheet.cell(row, column)

        source_isin = _manual_text(cell(isin_column), "isin", required=True)
        assert source_isin is not None
        try:
            normalized_isin = normalize_isin(source_isin)
        except InvalidIsinError as exc:
            raise _RowValidationError("isin", exc.message, source_isin) from exc
        target_event = _event_type(_manual_text(cell(4), "target_event_type", required=True))
        return ImportRow(
            row_number=row,
            source_isin=source_isin,
            normalized_isin=normalized_isin,
            source_name=_text(cell(3).value),
            target_event_type=target_event,
            purchase_date=_manual_date(cell(5), "purchase_date"),
            target_event_date=_manual_date(cell(7), "target_event_date"),
            quantity=_manual_decimal(cell(8), "quantity", positive=True),
            purchase_clean_price_rub_per_bond=_manual_decimal(
                cell(9), "purchase_clean_price_rub_per_bond", non_negative=True
            ),
            purchase_accrued_interest_rub_per_bond=_manual_decimal(
                cell(10), "purchase_accrued_interest_rub_per_bond", non_negative=True
            ),
            purchase_commission_rub_per_bond=_manual_decimal(
                cell(12), "purchase_commission_rub_per_bond", non_negative=True
            ),
            target_redemption_price_rub_per_bond=_optional_decimal(cell(14)),
            sale_commission_rub_per_bond=_optional_decimal(cell(18)) or Decimal(0),
            planned_yield_manual_reference=_optional_decimal(cell(22)),
            offer_submission_period=_text(cell(6).value),
        )

    @staticmethod
    def _is_empty_row(cells: Sequence[SpreadsheetCell]) -> bool:
        return all(cell.value is None or str(cell.value).strip() == "" for cell in cells)


class XlsxPortfolioReader(SpreadsheetPortfolioReader):
    """Backward-compatible name for the multi-format spreadsheet reader."""


class _RowValidationError(ValueError):
    def __init__(self, field: str, message: str, source_value: str | None = None) -> None:
        super().__init__(message)
        self.field = field
        self.message = message
        self.source_value = source_value


def _reject_formula(cell: SpreadsheetCell, field: str) -> None:
    if cell.is_formula or (isinstance(cell.value, str) and cell.value.startswith("=")):
        raise _RowValidationError(
            field, "A manual source field cannot contain a formula", str(cell.value)
        )


def _manual_text(cell: SpreadsheetCell, field: str, *, required: bool) -> str | None:
    _reject_formula(cell, field)
    value = _text(cell.value)
    if required and value is None:
        raise _RowValidationError(field, "A required value is missing")
    return value


def _manual_date(cell: SpreadsheetCell, field: str) -> date:
    _reject_formula(cell, field)
    value = cell.value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError as exc:
            raise _RowValidationError(field, "Date must use YYYY-MM-DD", value) from exc
    raise _RowValidationError(field, "A valid date is required", _text(value))


def _manual_decimal(
    cell: SpreadsheetCell,
    field: str,
    *,
    positive: bool = False,
    non_negative: bool = False,
) -> Decimal:
    _reject_formula(cell, field)
    value = _decimal(cell.value, field)
    if positive and value <= 0:
        raise _RowValidationError(field, "Value must be positive", str(cell.value))
    if non_negative and value < 0:
        raise _RowValidationError(field, "Value cannot be negative", str(cell.value))
    return value


def _optional_decimal(cell: SpreadsheetCell) -> Decimal | None:
    if cell.value in (None, "") or cell.is_formula:
        return None
    return _decimal(cell.value, "optional_decimal")


def _decimal(value: Any, field: str) -> Decimal:
    if value in (None, ""):
        raise _RowValidationError(field, "A numeric value is required")
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except InvalidOperation as exc:
        raise _RowValidationError(field, "A valid decimal value is required", str(value)) from exc


def _event_type(value: str | None) -> str:
    normalized = (value or "").strip().lower()
    if normalized in {"погашение", "maturity"}:
        return "maturity"
    if normalized in {"оферта", "offer"}:
        return "offer"
    raise _RowValidationError(
        "target_event_type", "Event type must be maturity/погашение or offer/оферта", value
    )


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
